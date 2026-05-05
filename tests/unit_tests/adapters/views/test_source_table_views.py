from io import BytesIO

from openpyxl import load_workbook
import pytest

from model_builder.adapters.repositories import SessionSystemRepository
from model_builder.domain.entities.web_core.model_web import ModelWeb


def _setup_session(client, system_data: dict) -> None:
    SessionSystemRepository(client.session).save_data(system_data)


def _first_editable_source_row(client):
    model_web = ModelWeb(SessionSystemRepository(client.session))
    return next(eq for eq in model_web.web_explainable_quantities_sources if not eq.is_calculated)


@pytest.mark.django_db
class TestSourceTableViews:

    def test_source_table_does_not_prerender_row_editor_forms(self, client, minimal_system_data):
        _setup_session(client, minimal_system_data)

        response = client.get("/model_builder/source-table/")

        body = response.content.decode()
        assert response.status_code == 200
        assert "source-table-row-editor/" in body
        assert 'data-action="source-table-row-edit"' not in body
        assert "source-editor-select" not in body
        assert 'id="source-table-confidence-menu-template"' in body
        assert 'id="source-cell-' in body
        assert 'id="comment-cell-' in body
        assert body.count('class="confidence-menu"') == 1

    def test_source_table_row_editor_renders_one_edit_form(self, client, minimal_system_data, monkeypatch):
        _setup_session(client, minimal_system_data)
        eq = _first_editable_source_row(client)
        object_id = eq.modeling_obj_container.efootprint_id
        attr_name = eq.attr_name_in_mod_obj_container
        monkeypatch.setattr(
            "model_builder.adapters.views.views.ModelWeb",
            lambda *args, **kwargs: pytest.fail("source_table_row_editor should not hydrate ModelWeb"),
        )

        response = client.get(f"/model_builder/source-table-row-editor/{object_id}/{attr_name}/")

        body = response.content.decode()
        field_name_prefix = f"{eq.modeling_obj_container.class_as_simple_str}_{attr_name}"
        assert response.status_code == 200
        assert 'data-action="source-table-row-edit"' in body
        assert "source-editor-select" in body
        assert 'hx-trigger="source-table-row-edit-submit"' in body
        assert f'data-source-row-id="{eq.web_id}"' in body
        assert "data-source-table-url" not in body
        assert f'name="{field_name_prefix}__source_id"' in body
        assert f'name="{field_name_prefix}__comment"' in body

    def test_source_table_row_editor_rejects_non_source_row_attribute(self, client, minimal_system_data):
        _setup_session(client, minimal_system_data)
        eq = _first_editable_source_row(client)
        object_id = eq.modeling_obj_container.efootprint_id

        response = client.get(f"/model_builder/source-table-row-editor/{object_id}/name/")

        assert response.status_code == 404

    def test_source_table_row_editor_rejects_non_source_attribute_even_when_json_looks_source_like(
            self, client, minimal_system_data):
        _setup_session(client, minimal_system_data)
        eq = _first_editable_source_row(client)
        object_id = eq.modeling_obj_container.efootprint_id
        object_type = eq.modeling_obj_container.class_as_simple_str
        system_data = SessionSystemRepository(client.session).get_system_data()
        system_data[object_type][object_id]["name"] = {
            "label": "Name",
            "source": eq.source.id,
            "comment": "forged source metadata",
        }
        _setup_session(client, system_data)

        response = client.get(f"/model_builder/source-table-row-editor/{object_id}/name/")

        assert response.status_code == 404

    def test_download_sources_exports_confidence_and_comment_columns(self, client, minimal_system_data):
        _setup_session(client, minimal_system_data)
        model_web = ModelWeb(SessionSystemRepository(client.session))
        eq = next(row for row in model_web.web_explainable_quantities_sources if not row.is_calculated)
        eq.efootprint_object.confidence = "high"
        eq.efootprint_object.comment = "exported metadata note"
        model_web.persist_to_cache()

        response = client.get("/model_builder/download-sources/")

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook["Sources"]
        rows = list(worksheet.iter_rows(values_only=True))
        assert response.status_code == 200
        assert rows[0] == (
            "Item name",
            "Attribute of",
            "Object type",
            "Value",
            "Unit",
            "Source name",
            "Source link",
            "confidence",
            "comment",
        )
        assert [row[-2:] for row in rows[1:]].count(("high", "exported metadata note")) == 1
